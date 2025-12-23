# modules/excel_export.py
"""
Excel export tool voor het exporteren van video's en prestaties.
Gebruikt openpyxl voor Excel-bestandenwerk.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from modules.video_DataAccess import videos_ophalen
from modules.prestatie_DataAccess import prestaties_ophalen


def _format_header(ws, row_num):
    """Formatteer de header-rij met vet lettertype en achtergrondkleur."""
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[row_num]:
        if cell.value:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border


def exporteer_videos_excel(bestandsnaam=None):
    """
    Exporteert alle videos naar een Excel bestand.
    
    Args:
        bestandsnaam (str): Optioneel. Naam van het Excel-bestand.
                           Default: 'export_videos_YYYY-MM-DD.xlsx'
    
    Returns:
        str: Naam van het aangemaakte bestand.
    """
    if bestandsnaam is None:
        datum = datetime.now().strftime("%Y-%m-%d")
        bestandsnaam = f"export_videos_{datum}.xlsx"
    
    # Haal alle videos op (nu Video objecten!)
    videos = videos_ophalen()
    
    # Maak nieuw Excel-werkboek
    wb = Workbook()
    ws = wb.active
    ws.title = "Videos"
    
    # Header rij
    headers = ["ID", "Titel", "Platform", "Status", "Datum Aangemaakt", "Datum Gepost"]
    ws.append(headers)
    _format_header(ws, 1)
    
    # Voeg video-data toe
    for video in videos:
        ws.append([
            video.id,
            video.titel,
            video.platform,
            video.status,
            video.datum_aangemaakt,
            video.datum_gepost if video.datum_gepost else "-"
        ])
    
    # Breid kolombreedte automatisch uit
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Bewaar bestand
    wb.save(bestandsnaam)
    return bestandsnaam


def exporteer_prestaties_excel(bestandsnaam=None):
    """
    Exporteert alle prestaties (metrics) naar een Excel bestand.
    
    Args:
        bestandsnaam (str): Optioneel. Naam van het Excel-bestand.
                           Default: 'export_prestaties_YYYY-MM-DD.xlsx'
    
    Returns:
        str: Naam van het aangemaakte bestand.
    """
    if bestandsnaam is None:
        datum = datetime.now().strftime("%Y-%m-%d")
        bestandsnaam = f"export_prestaties_{datum}.xlsx"
    
    # Haal alle prestaties op (nu Prestatie objecten!)
    prestaties = prestaties_ophalen()
    
    # Maak nieuw Excel-werkboek
    wb = Workbook()
    ws = wb.active
    ws.title = "Prestaties"
    
    # Header rij
    headers = ["ID", "Video Titel", "Datum Gemeten", "Views", "Likes", "Comments", "Shares"]
    ws.append(headers)
    _format_header(ws, 1)
    
    # Voeg prestatie-data toe
    for prestatie in prestaties:
        ws.append([
            prestatie.id,
            prestatie.video_titel,
            prestatie.datum_gemeten,
            prestatie.views,
            prestatie.likes,
            prestatie.comments,
            prestatie.shares
        ])
    
    # Breid kolombreedte automatisch uit
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Voeg aantal rijen toe in sheet naam
    wb.save(bestandsnaam)
    return bestandsnaam


def exporteer_alles_excel(bestandsnaam=None):
    """
    Exporteert ALLES (videos en prestaties) naar één Excel bestand met 2 sheets.
    
    Args:
        bestandsnaam (str): Optioneel. Naam van het Excel-bestand.
                           Default: 'export_compleet_YYYY-MM-DD.xlsx'
    
    Returns:
        str: Naam van het aangemaakte bestand.
    """
    if bestandsnaam is None:
        datum = datetime.now().strftime("%Y-%m-%d")
        bestandsnaam = f"export_compleet_{datum}.xlsx"
    
    videos = videos_ophalen()
    prestaties = prestaties_ophalen()
    
    # Maak werkboek met 2 sheets
    wb = Workbook()
    wb.remove(wb.active)  # Verwijder de default lege sheet
    
    # SHEET 1: Videos
    ws_videos = wb.create_sheet("Videos")
    headers_videos = ["ID", "Titel", "Platform", "Status", "Datum Aangemaakt", "Datum Gepost"]
    ws_videos.append(headers_videos)
    _format_header(ws_videos, 1)
    
    for video in videos:
        ws_videos.append([
            video.id,
            video.titel,
            video.platform,
            video.status,
            video.datum_aangemaakt,
            video.datum_gepost if video.datum_gepost else "-"
        ])
    
    # Kolombreedte Videos
    for column in ws_videos.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_videos.column_dimensions[column_letter].width = max_length + 2
    
    # SHEET 2: Prestaties
    ws_prestaties = wb.create_sheet("Prestaties")
    headers_prestaties = ["ID", "Video Titel", "Datum Gemeten", "Views", "Likes", "Comments", "Shares"]
    ws_prestaties.append(headers_prestaties)
    _format_header(ws_prestaties, 1)
    
    for prestatie in prestaties:
        ws_prestaties.append([
            prestatie.id,
            prestatie.video_titel,
            prestatie.datum_gemeten,
            prestatie.views,
            prestatie.likes,
            prestatie.comments,
            prestatie.shares
        ])
    
    # Kolombreedte Prestaties
    for column in ws_prestaties.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_prestaties.column_dimensions[column_letter].width = max_length + 2
    
    wb.save(bestandsnaam)
    return bestandsnaam
